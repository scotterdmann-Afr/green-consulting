import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import shutil, os, json, sys

def generate_devis(data, template_path, output_path):
    """
    data = {
      "date": "2026-03-17",
      "numero": "DE202603002",
      "client": "SOBRAGA",
      "categorie": "Pdr Garage",
      "sous_categorie": "REGULARISATION",  # optional
      "items": [
        {"designation": "Injecteur 23260-UB020", "quantite": 20, "prix_unitaire": 162200}
      ],
      "livraison": "30 jours"  # optional
    }
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['Facture']

    # Date
    if data.get('date'):
        ws['D4'] = datetime.strptime(data['date'], '%Y-%m-%d')
    
    # Devis N°
    ws['D5'] = data.get('numero', '')
    
    # Client
    ws['C10'] = data.get('client', '')
    
    # Sous-catégorie / catégorie
    row = 17
    if data.get('sous_categorie'):
        ws.cell(row=row, column=1).value = data['sous_categorie']
        row += 1
    if data.get('categorie'):
        ws.cell(row=row, column=1).value = data['categorie']
        row += 1
    
    # Items - fill from row 'row' up to max row 37
    for item in data.get('items', []):
        if row > 37:
            break
        ws.cell(row=row, column=1).value = item.get('designation', '')
        ws.cell(row=row, column=2).value = float(item.get('quantite', 0))
        ws.cell(row=row, column=3).value = float(item.get('prix_unitaire', 0))
        ws.cell(row=row, column=4).value = f'=C{row}*B{row}'
        row += 1
    
    # Livraison
    if data.get('livraison'):
        ws['A51'] = data['livraison']
    
    wb.save(output_path)
    return output_path


def generate_facture(data, template_path, output_path):
    """
    data = {
      "date": "2026-02-20",
      "numero": "FAC202602004",
      "client": "SOBRAGA\nNIF 790 037 F",
      "da": "DA 266947",
      "categorie": "Pour Invest Etiqueteuse L03",
      "items": [...],
      "modalites": "Le solde à la livraison"
    }
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['Facture']
    
    if data.get('date'):
        ws['D4'] = datetime.strptime(data['date'], '%Y-%m-%d')
    ws['D5'] = data.get('numero', '')
    ws['C10'] = data.get('client', '')
    
    if data.get('da'):
        ws['C16'] = data['da']
    
    row = 19
    if data.get('categorie'):
        ws.cell(row=row, column=1).value = data['categorie']
        row += 1
        row += 1  # blank row like in example (row 21 empty)
    
    for item in data.get('items', []):
        if row > 49:
            break
        ws.cell(row=row, column=1).value = item.get('designation', '')
        ws.cell(row=row, column=2).value = float(item.get('quantite', 0))
        ws.cell(row=row, column=3).value = float(item.get('prix_unitaire', 0))
        ws.cell(row=row, column=4).value = f'=C{row}*B{row}'
        row += 1
    
    if data.get('modalites'):
        ws['A59'] = f"Modalités de paiement :\n{data['modalites']}"
    
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    # Test with SOBRAGA devis data
    test_devis = {
        "date": "2026-03-17",
        "numero": "DE202603002",
        "client": "SOBRAGA",
        "sous_categorie": "REGULARISATION",
        "categorie": "Pdr Garage",
        "items": [
            {"designation": "Injecteur 23260-UB020", "quantite": 20, "prix_unitaire": 162200}
        ],
        "livraison": "-"
    }
    
    test_facture = {
        "date": "2026-02-20",
        "numero": "FAC202602004",
        "client": "SOBRAGA\nNIF 790 037 F",
        "da": "DA 266947",
        "categorie": "Pour Invest Etiqueteuse L03",
        "items": [
            {"designation": "SIMATIC Memory Card 256 MB 6ES7 954-8LL04-0AA0", "quantite": 2, "prix_unitaire": 348000}
        ],
        "modalites": "Le solde à la livraison"
    }
    
    out_devis = '/home/claude/TEST_DE202603002_SOBRAGA.xlsx'
    out_facture = '/home/claude/TEST_FA202602004_SOBRAGA.xlsx'
    
    generate_devis(test_devis, '/mnt/user-data/uploads/template_devis_vierge.xlsx', out_devis)
    generate_facture(test_facture, '/mnt/user-data/uploads/Template_facture_vierge.xlsx', out_facture)
    
    print(f"Devis generated: {out_devis}")
    print(f"Facture generated: {out_facture}")
